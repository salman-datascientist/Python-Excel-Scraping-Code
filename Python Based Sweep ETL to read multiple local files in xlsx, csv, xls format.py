# Databricks notebook source
# MAGIC %md ## Notes for running this script
# MAGIC
# MAGIC
# MAGIC * Re-Save .CSV files as .XLSX files format (script does not handle .CSV files)
# MAGIC * Re-Save .xls (old excel format) to new .XLSX file format
# MAGIC * First Row in every sheet/worktab should always be a header
# MAGIC * Sensitive information should be reviewed and masked (for changes in the code need to be eplicitly mentioned)

# COMMAND ----------

# MAGIC %md ## Files and Folders Cleanup Maintanence Code

# COMMAND ----------

#folder_path = 'insert here'

#try:
#    #os.rmdir(folder_path)
#    shutil.rmtree(folder_path)
#    print(f"Successfully removed empty folder: {folder_path}")
#except OSError as e:
#    print(f"Error: {e}")

# COMMAND ----------

# MAGIC %md ##Email Campaigns published on sharepoint
# MAGIC
# MAGIC https://towardsdatascience.com/cleansing-and-transforming-schema-drifted-csv-files-into-relational-data-in-azure-databricks-519e82ea84ff
# MAGIC
# MAGIC
# MAGIC https://adb-1411802526182681.1.azuredatabricks.net/?o=1411802526182681#notebook/696832424309190/command/4102857815304094

# COMMAND ----------

# MAGIC %md ##  Beginners guide for text preprocessing in NLP
# MAGIC
# MAGIC * https://swatimeena989.medium.com/beginners-guide-for-preprocessing-text-data-f3156bec85ca

# COMMAND ----------

# MAGIC %md ## Import Python Libraries

# COMMAND ----------

# MAGIC %pip install rapidfuzz --quiet --ignore-installed

# COMMAND ----------

# MAGIC %pip install xlrd --quiet --ignore-installed

# COMMAND ----------

import os
import glob
import openpyxl
import xlrd
import shutil
import datetime
import re              # regular expressions
import rapidfuzz
import pandas as pd
import numpy as np
# Set the display option to show all columns
pd.set_option('display.max_columns', None)

# Set the display option to show the full content of 'LongString' column
pd.set_option('display.max_colwidth', None)

# COMMAND ----------

# MAGIC %md ## Import Python Libraries for NLP tasks

# COMMAND ----------

import spacy
import nltk
from nltk.corpus import stopwords
import unicodedata
import string

# COMMAND ----------

# MAGIC %md ## Set Default folder for reading and moving files

# COMMAND ----------

# Set the working directory to the directory containing the list of Excel files
os.chdir('insert here')

# COMMAND ----------

# MAGIC %md ## Source Files Location where files new files are uploaded

# COMMAND ----------

# Define the folder path where you want to search for Excel files
folder_path = "insert here" # Replace with the actual destination directory path with forward slash in the end
print(folder_path)

# COMMAND ----------

# MAGIC %md ## Error Files Location where files are moved when Error occured reading files

# COMMAND ----------

# Specify the destination directory where you want to move the files if libraries are not able to read or open the files
error_folder_path = "insert here"  # Replace with the actual destination directory path with forward slash in the end
print(error_folder_path)

# COMMAND ----------

# MAGIC %md ## Successful Files Location where files are moved after Successful Read of file

# COMMAND ----------

# Specify the destination directory where you want to move the files when libraries opened and read the files completely
successful_read_files_foder_path = "insert here"  # Replace with the actual destination directory path with forward slash in the end
print(successful_read_files_foder_path)

# COMMAND ----------

# MAGIC %md ## New Folder Creation using Date Time Stamp in Error and Successful folders
# MAGIC
# MAGIC * Check new files are there in inbox folder
# MAGIC * Create a new folder with current date time stamp
# MAGIC * Move the files to the new folder with current date time stamp - to check how many files were read and check the audit log

# COMMAND ----------

# Get a list of all files in the folder
file_list = os.listdir(folder_path)

# Initialize a variable to track if there are any files in the folder
excel_files_exist = False

# COMMAND ----------

# Define a list of valid Excel file extensions
excel_extensions = ['.xlsx', '.xls', '.csv']

# Iterate through the items to check if any of them are files
for file_name in file_list:
  # Check if the file extension is in the list of valid Excel extensions
  if any(file_name.endswith(ext) for ext in excel_extensions):
    excel_files_exist = True
    break  # Exit the loop as soon as an Excel file is found

# Check if excel files exist in the folder
if excel_files_exist:
  print("There are new files in the folder place_source_files_inbox.")
else:
  print("The folder place_source_files_inbox is empty or contains only subdirectories.")

# COMMAND ----------

if excel_files_exist:

    # Get the current date and time
    current_datetime = datetime.datetime.now()

    # Format the current date and time as a string (e.g., "2023-10-26_145855" for October 26, 2023, 14:58:55)
    timestamp = current_datetime.strftime("%Y-%m-%d_%H%M%S")

    # Create the new folder with the timestamp
    new_folder_name_error = os.path.join(error_folder_path, timestamp)

    try:
        os.mkdir(new_folder_name_error)
        print(f"Created folder in Error Folder: {new_folder_name_error}")
    except OSError as e:
        print(f"Failed to create folder in Error Folder: {new_folder_name_error}")
        print(f"Error: {e}")

    # Create the new folder with the timestamp
    new_folder_name_success = os.path.join(successful_read_files_foder_path, timestamp)

    try:
        os.mkdir(new_folder_name_success)
        print(f"Created folder in Successful Read Folder: {new_folder_name_success}")
    except OSError as e:
        print(f"Failed to create folder in Successful Read Folder: {new_folder_name_success}")
        print(f"Error: {e}")
else:
    print("New folders not created as source folder is empty.")

# COMMAND ----------

# MAGIC %md ## Step 1: Script to read all .xls, .xlsx and .csv files from the location and keep only the files that can be read properly
# MAGIC
# MAGIC * Step 1: If read file lead to success then move to success folder
# MAGIC * Step 2: If read file lead to error then move to error folder

# COMMAND ----------

if excel_files_exist:
    # Use the glob.glob() function to find all Excel files in the folder
    excel_files = glob.glob(os.path.join(folder_path, "*.xlsx")) + glob.glob(os.path.join(folder_path, "*.xls")) + glob.glob(os.path.join(folder_path, "*.csv"))

    i = 0

    for excel_file_path in excel_files:

        file_name_with_extension = os.path.basename(excel_file_path)
        file_full_path = folder_path + file_name_with_extension

        # Check the file extension as previous code did not run correctly
        file_extension = os.path.splitext(excel_file_path)[1].lower()
        print(file_extension)
        print(file_name_with_extension)
        print(file_full_path)
    
        if file_extension == ".xls":
            # Handle .xls files using xlrd
            print("This is an XLS (Excel 2003) file. Attempting to Read the file")
            try:
                workbook_xlrd = xlrd.open_workbook(file_full_path)
                i = i+1 
                # Check the number of worksheets using xlrd
                sheet_count_xlrd = len(workbook_xlrd.sheet_names())
        
                if sheet_count_xlrd > 1:
                    # Iterate through all sheets in the workbook_xlrd
                    #for sheet in workbook.sheets():
                    for sheet_name in workbook_xlrd.sheet_names():
                        sheet = workbook_xlrd.sheet_by_name(sheet_name)
                        max_row = sheet.nrows
                        max_column = sheet.ncols
                        print(f"The Excel file '{file_full_path}' (using xlrd) has {sheet_count_xlrd} worksheets and {sheet_name} has {max_row} rows and {max_column} columns.")
                elif sheet_count_xlrd == 1:
                    # Iterate through the single sheet in the workbook_xlrd
                    #for sheet in workbook.sheets():
                    for sheet_name in workbook_xlrd.sheet_names():
                        sheet = workbook_xlrd.sheet_by_name(sheet_name)
                        max_row = sheet.nrows
                        max_column = sheet.ncols
                        print(f"The Excel file '{file_full_path}' (using xlrd) has only one worksheet and has {max_row} rows and {max_column} columns.")
                else:
                    print(f"The Excel file '{file_full_path}' (using xlrd) is empty.")
                
                # Construct the full path to the destination file in the new directory
                destination_file_path = os.path.join(new_folder_name_success, file_name_with_extension)
                shutil.move(file_full_path, destination_file_path)
                print(f"File '{file_name_with_extension}' moved to '{new_folder_name_success}'")
            except xlrd.XLRDError as e:
                print(f"An error occurred while trying to read the Excel file: {e}")
                destination_file_path = os.path.join(new_folder_name_error, file_name_with_extension)
                # Move the file to the destination directory
                shutil.move(file_full_path, destination_file_path)
                print(f"File '{file_name_with_extension}' moved to '{new_folder_name_error}'")     
            except FileNotFoundError:    
                print(f"The file '{file_full_path}' was not found.")
                destination_file_path = os.path.join(new_folder_name_error, file_name_with_extension)
                # Move the file to the destination directory
                shutil.move(file_full_path, destination_file_path)
                print(f"File '{file_name_with_extension}' moved to '{new_folder_name_error}'")          
            except Exception as e:
                print(f"An error occurred: {e}")
                # Construct the full path to the destination file in the new directory
                destination_file_path = os.path.join(new_folder_name_error, file_name_with_extension)
                # Move the file to the destination directory
                shutil.move(file_full_path, destination_file_path)
                print(f"File '{file_name_with_extension}' moved to '{new_folder_name_error}'")
            finally:
                # Close the workbook if it was successfully opened
                if 'workbook_xlrd' in locals():
                    workbook_xlrd.release_resources()        
        elif file_extension == ".xlsx":
            # Handle .xlsx files using openpyxl
            print("This is an XLSX (Excel 2007 or later) file. Attempting to Read the file")
            # Load the Excel workbook in .xlsx format
            try:
                workbook = openpyxl.load_workbook(file_full_path, data_only=True, read_only=True) #important condition to ignore formatting of data
                i = i+1
                # Check the number of worksheets
                sheet_count = len(workbook.sheetnames)
                if sheet_count > 1:
                    # Iterate through all sheets in the workbook
                    for sheet in workbook:
                        max_row = sheet.max_row
                        max_column = sheet.max_column
                        sheet_name = sheet.title
                        print(f"The Excel file '{file_full_path}' (using openxl) has {sheet_count} worksheets and {sheet_name} has {max_row} rows and {max_column} columns.")
                elif sheet_count == 1:
                    for sheet in workbook:
                        max_row = sheet.max_row
                        max_column = sheet.max_column
                        sheet_name = sheet.title
                        print(f"The Excel file '{file_full_path}' (using openxl)  has only one worksheet and has {max_row} rows and {max_column} columns.")
                else:
                    print(f"The Excel file '{file_full_path}' (using openxl) is empty.")

                # Close the workbook
                workbook.close()
                # Construct the full path to the destination file in the new directory
                destination_file_path = os.path.join(new_folder_name_success, file_name_with_extension)
                shutil.move(file_full_path, destination_file_path)
                print(f"File '{file_name_with_extension}' moved to '{new_folder_name_success}'")
            except FileNotFoundError:
                print(f"The file '{file_full_path}' was not found.")
                # Construct the full path to the destination file in the new directory
                destination_file_path = os.path.join(new_folder_name_error, file_name_with_extension)
                # Move the file to the destination directory
                shutil.move(file_full_path, destination_file_path)          
                print(f"File '{file_name_with_extension}' moved to '{new_folder_name_error}'")
            except Exception as e:
                print(f"An error occurred: {e}")
                # Construct the full path to the destination file in the new directory
                destination_file_path = os.path.join(new_folder_name_error, file_name_with_extension)
                # Move the file to the destination directory
                shutil.move(file_full_path, destination_file_path)          
                print(f"File '{file_name_with_extension}' moved to '{new_folder_name_error}'")
            finally:
                #Close the workbook
                workbook.close()
        elif file_extension == ".csv":
            # Handle .csv files using read_csv
            print("This is an CSV file. Attempting to Read the file")
            try:
                #csv_file = pd.read_csv(file_full_path, data_only=True, read_only=True) #important condition to ignore formatting of data 
                csv_file = pd.read_csv(file_full_path) #important condition to ignore formatting of data 
                i = i + 1
                # Getting the number of rows and columns
                num_rows, num_columns = csv_file.shape
                if num_rows > 1 and num_columns > 1:
                    print(f"The CSV file '{file_full_path}' (using read_csv) has {num_rows} rows and {num_columns} columns.")
                else:
                    print(f"The CSV file '{file_full_path}' (using read_csv) is empty.")
                # Construct the full path to the destination file in the new directory
                destination_file_path = os.path.join(new_folder_name_success, file_name_with_extension)
                shutil.move(file_full_path, destination_file_path)
                print(f"File '{file_name_with_extension}' moved to '{new_folder_name_success}'")
            except FileNotFoundError:
                print(f"The file '{file_full_path}' was not found.")
                # Construct the full path to the destination file in the new directory
                destination_file_path = os.path.join(new_folder_name_error, file_name_with_extension)
                # Move the file to the destination directory
                shutil.move(file_full_path, destination_file_path)          
                print(f"File '{file_name_with_extension}' moved to '{new_folder_name_error}'")
            except Exception as e:
                print(f"An error occurred: {e}")
                # Construct the full path to the destination file in the new directory
                destination_file_path = os.path.join(new_folder_name_error, file_name_with_extension)
                # Move the file to the destination directory
                shutil.move(file_full_path, destination_file_path)          
                print(f"File '{file_name_with_extension}' moved to '{new_folder_name_error}'")                
        else:
            # Handle other file types or show an error message
            print("This is neither a CSV nor an XLS nor an XLSX file.")
            # Construct the full path to the destination file in the new directory
            destination_file_path = os.path.join(new_folder_name_error, file_name_with_extension)
            # Move the file to the destination directory
            shutil.move(file_full_path, destination_file_path)          
            print(f"File '{file_name_with_extension}' moved to '{new_folder_name_error}'")  

    print(f"The total number of files successfully read is '{i}' files.")
else:
    print("Read File Logic was not run as source folder is empty.")

# COMMAND ----------

# MAGIC %md ## Step 2: Read all the good excel files into pandas dataframe

# COMMAND ----------

print(new_folder_name_success)
print(new_folder_name_error)

# COMMAND ----------

# MAGIC %md ## Step 3: Read all the good excel files into pandas dataframe dictionary

# COMMAND ----------

# MAGIC %md ### ETL Bronze Layer

# COMMAND ----------

# Step 1 : Use the glob.glob() function to find all Excel files in the folder
excel_files = glob.glob(os.path.join(new_folder_name_success, "*.xlsx")) + glob.glob(os.path.join(new_folder_name_success, "*.xls")) + glob.glob(os.path.join(new_folder_name_success, "*.csv"))

# Partial Column Names  (do not delete)
partial_column_names = ["MobileNumber", "FirstName", "MiddleName", "LastName", "Addr1", "Addr2", "City", "State", "Zip", "OptOutURL", "Var1", "Var2", "Var3", "Var4", "CID", "SubscriberKey", "CreatedDate", "LastModifiedDate", "CampaignName", "SuppressionType", "CampaignRunID", "HighLowIP", "Sent", "sentdate", "Open", "open date", "click", "clickdate", "Hardbounce", "softbounce", "blockedbounce", "unsub"]

# List of strings to remove
strings_to_remove = [".xlsx", ".xls", ".csv"]

# Create a regular expression pattern by joining the strings with the "|" (OR) operator
pattern = '|'.join(map(re.escape, strings_to_remove))

# Create an empty dictionary to store Bronze DataFrames
bronze_dataframes = {}

i = 0

# Step 2: Iterate through the files
for excel_file_path in excel_files:

  file_name_with_extension = os.path.basename(excel_file_path)
  file_full_path = new_folder_name_success + file_name_with_extension

  # Check the file extension as previous code did not run correctly
  file_extension = os.path.splitext(excel_file_path)[1].lower()
  print(file_extension)
  print(file_name_with_extension)
  print(file_full_path)
  # Use re.sub() to remove the strings
  file_name_without_extension = re.sub(pattern, '', file_name_with_extension)
    
  if file_extension == ".xls":
    # Handle .xls files using xlrd
    print("This is an XLS (Excel 2003) file. Attempting to Read the file")
    workbook_xlrd = xlrd.open_workbook(file_full_path)
    # Check the number of worksheets using xlrd
    sheet_count_xlrd = len(workbook_xlrd.sheet_names())

    if sheet_count_xlrd > 1:
      # Iterate through all sheets in the workbook_xlrd
      for sheet_name in workbook_xlrd.sheet_names():
        sheet = workbook_xlrd.sheet_by_name(sheet_name)
        max_row = sheet.nrows
        max_column = sheet.ncols
        print(f"'{file_full_path}' has '{sheet}' with '{max_row}' rows and '{max_column}' columns.")
        if max_column > 1 and max_row > 1:
          try:
            # Read all worksheets into a dictionary of DataFrames
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=0, engine='xlrd', dtype=str)

            # Step 4: Check for partial column name matches
            matching_columns = [col for col in df.columns if any(partial in col for partial in partial_column_names)]

            if matching_columns:
                # Step 5: Store the DataFrame with file name and worksheet name as keys
                key = f"{file_name_without_extension}_{sheet_name}"
                bronze_dataframes[key] = df
          
          except Exception as e:
            print(f"An error occurred: {e}")

    elif sheet_count_xlrd == 1:
      # Iterate through all sheets in the workbook_xlrd
      for sheet_name in workbook_xlrd.sheet_names():
        sheet = workbook_xlrd.sheet_by_name(sheet_name)
        max_row = sheet.nrows
        max_column = sheet.ncols
        print(f"'{file_full_path}' has '{sheet}' with '{max_row}' rows and '{max_column}' columns.")
        if max_column > 1 and max_row > 1:
          try:
            # Read all worksheets into a dictionary of DataFrames
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=0, engine='xlrd', dtype=str)

            # Step 4: Check for partial column name matches
            matching_columns = [col for col in df.columns if any(partial in col for partial in partial_column_names)]

            if matching_columns:
              # Step 5: Store the DataFrame with file name and worksheet name as keys
              key = f"{file_name_without_extension}_{sheet_name}"
              bronze_dataframes[key] = df
          
          except Exception as e:
            print(f"An error occurred: {e}")                 

    # Close the workbook if it was successfully opened
    if 'workbook_xlrd' in locals():
      workbook_xlrd.release_resources()  

  elif file_extension == ".xlsx":
    # Handle .xlsx files using openpyxl
    print("This is an XLSX (Excel 2007 or later) file. Attempting to Read the file")
    workbook_xlsx = openpyxl.load_workbook(file_full_path, data_only=True, read_only=True)           #important condition to ignore formatting of data
    # Check the number of worksheets
    sheet_count_xlsx = len(workbook_xlsx.sheetnames)

    if sheet_count_xlsx > 1:
      # Iterate through all sheets in the workbook_xlsx  
      for sheet in workbook_xlsx:
        sheet_name = sheet.title
        max_row = sheet.max_row
        max_column = sheet.max_column
        print(f"'{file_full_path}' has '{sheet}' with '{max_row}' rows and '{max_column}' columns.")
        if max_column > 1 and max_row > 1:
          try:
            # Read all worksheets into a dictionary of DataFrames
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=0, engine='openpyxl', dtype=str)

            # Step 4: Check for partial column name matches
            matching_columns = [col for col in df.columns if any(partial in col for partial in partial_column_names)]

            if matching_columns:
              # Step 5: Store the DataFrame with file name and worksheet name as keys
              key = f"{file_name_without_extension}_{sheet_name}"
              bronze_dataframes[key] = df
          
          except Exception as e:
            print(f"An error occurred: {e}")

    elif sheet_count_xlsx == 1:
      # Iterate through all sheets in the workbook_xlsx
      for sheet in workbook_xlsx:
        sheet_name = sheet.title
        max_row = sheet.max_row
        max_column = sheet.max_column
        print(f"'{file_full_path}' has '{sheet}' with '{max_row}' rows and '{max_column}' columns.")
        if max_column > 1 and max_row > 1:
          try:
            # Read all worksheets into a dictionary of DataFrames
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=0, engine='openpyxl', dtype=str)

            # Step 4: Check for partial column name matches
            matching_columns = [col for col in df.columns if any(partial in col for partial in partial_column_names)]

            if matching_columns:
              # Step 5: Store the DataFrame with file name and worksheet name as keys
              key = f"{file_name_without_extension}_{sheet_name}"
              bronze_dataframes[key] = df
          except Exception as e:
            print(f"An error occurred: {e}")

  elif file_extension == ".csv":
    df = pd.read_csv(excel_file_path, header=0, index_col=False, dtype=str)

    # Get the number of rows and columns
    max_row, max_column = df.shape
    print(f"'{file_full_path}' has '{sheet}' with '{max_row}' rows and '{max_column}' columns.")
    if max_column > 1 and max_row > 1:
      try:
        # Step 4: Check for partial column name matches
        matching_columns = [col for col in df.columns if any(partial in col for partial in partial_column_names)]

        if matching_columns:
          # Step 5: Store the DataFrame with file name and worksheet name as keys
          key = f"{file_name_without_extension}"
          bronze_dataframes[key] = df
      except Exception as e:
        print(f"An error occurred: {e}")

# COMMAND ----------

# DBTITLE 1,Audit Step to check number of files read into a data dictionary called as bronze layer
# Get the keys as a view object
# keys_view = bronze_dataframes.keys()

# Convert the view object to a list if needed
# keys_list = list(keys_view)

# Print the keys
# print(keys_list)

# Print for better layout
# for key in bronze_dataframes.keys():
#    print(key)

# Iterate through the dictionary and check the data type of each value
for key, value in bronze_dataframes.items():
    data_type = type(value)
    print(f"Key: {key}, Data Type: {data_type}")

# COMMAND ----------

# Iterate through the dictionary of DataFrames
for key, df in bronze_dataframes.items():
    print(f"DataFrame Name: {key}")
    print(df.head(10))  # This will print the DataFrame
    print("=" * 40)  # Separator line for better readability

# COMMAND ----------

# MAGIC %md ### ETL Silver Layer
# MAGIC
# MAGIC Transforming and selecting the datacolumns that is requried to make into the table
# MAGIC
# MAGIC Looking for columns from each file
# MAGIC
# MAGIC * EmailAddress
# MAGIC * MobileNumber
# MAGIC * FirstName
# MAGIC * MiddleName
# MAGIC * LastName
# MAGIC * Addr1
# MAGIC * Addr2
# MAGIC * City
# MAGIC * State
# MAGIC * Zip
# MAGIC * OptOutURL
# MAGIC * Var1
# MAGIC * Var2
# MAGIC * Var3
# MAGIC * Var4
# MAGIC * CID
# MAGIC * SubscriberKey
# MAGIC * CreatedDate
# MAGIC * LastModifiedDate
# MAGIC * CampaignName
# MAGIC * SuppressionType
# MAGIC * CampaignRunID
# MAGIC * HighLowIP
# MAGIC * Sent
# MAGIC * sentdate
# MAGIC * Open
# MAGIC * opendate
# MAGIC * click
# MAGIC * clickdate
# MAGIC * Hardbounce
# MAGIC * softbounce
# MAGIC * blockedbounce
# MAGIC * unsub
# MAGIC

# COMMAND ----------

# DBTITLE 1,One time SQL to create empty table by the name sle_mun_data_silver_layer
# MAGIC %sql
# MAGIC
# MAGIC //*
# MAGIC drop table if exists sle_mun_data_silver_layer;
# MAGIC
# MAGIC -- Creates a Delta table
# MAGIC CREATE TABLE IF NOT EXISTS sle_mun_data_silver_layer ( rowid BIGINT NOT NULL,
# MAGIC emailaddress STRING,
# MAGIC mobilenumber STRING,
# MAGIC firstname STRING,
# MAGIC middlename STRING,
# MAGIC lastname STRING,
# MAGIC addr1 STRING,
# MAGIC addr2 STRING,
# MAGIC city STRING,
# MAGIC state STRING,
# MAGIC zip STRING,
# MAGIC optouturl STRING,
# MAGIC var1 STRING,
# MAGIC var2 STRING,
# MAGIC var3 STRING,
# MAGIC var4 STRING,
# MAGIC cid STRING,
# MAGIC subscriberkey STRING,
# MAGIC createddate STRING,
# MAGIC lastmodifieddate STRING,
# MAGIC campaignname STRING,
# MAGIC suppressiontype STRING,
# MAGIC campaignrunid STRING,
# MAGIC highlowip STRING,
# MAGIC sent STRING,
# MAGIC sentdate STRING,
# MAGIC open STRING,
# MAGIC opendate STRING,
# MAGIC click STRING,
# MAGIC clickdate STRING,
# MAGIC hardbounce STRING,
# MAGIC softbounce STRING,
# MAGIC blockedbounce STRING,
# MAGIC unsub STRING,
# MAGIC file_name STRING NOT NULL)
# MAGIC USING delta
# MAGIC PARTITIONED BY (file_name)
# MAGIC //*

# COMMAND ----------

# Step 1: Define search column names
search_column_names = ["EmailAddress", "MobileNumber", "FirstName", "MiddleName", "LastName", "Addr1", "Addr2", "City", "State", "Zip", "OptOutURL", "Var1", "Var2", "Var3", "Var4", "CID", "SubscriberKey", "CreatedDate", "LastModifiedDate", "CampaignName", "SuppressionType",
                        "CampaignRunID", "HighLowIP", "Sent", "sentdate", "Open", "open date", "click", "clickdate", "Hardbounce", "softbounce", "blockedbounce", "unsub"]

search_column_name_emailaddress = ['EmailAddress', 'emailaddress', 'EMAILADDRESS']
search_column_name_mobilenumber = ['MobileNumber', 'mobilenumber', 'MOBILENUMBER']
search_column_name_firstname = ['FirstName', 'firstname', 'FIRSTNAME']
search_column_name_middlename = ['MiddleName', 'middlename', 'MIDDLENAME']
search_column_name_lastname = ['LastName', 'lastname', 'LASTNAME']
search_column_name_addr1 = ['Addr1', 'addr1', 'ADDR1']
search_column_name_addr2 = ['Addr2', 'addr2', 'ADDR2']
search_column_name_city = ['City', 'city', 'CITY']
search_column_name_state = ['State', 'state', 'STATE']
search_column_name_zip = ['Zip', 'zip', 'ZIP']
search_column_name_optouturl = ['OptOutURL', 'optouturl', 'OPTOUTURL']
search_column_name_var1 = ['Var1', 'var1', 'VAR1']
search_column_name_var2 = ['Var2', 'var2', 'VAR2']
search_column_name_var3 = ['Var3', 'var3', 'VAR3']
search_column_name_var4 = ['Var4', 'var4', 'VAR4']
search_column_name_cid = ['CID', 'cid']
search_column_name_subscriberkey = ['SubscriberKey', 'subscriberkey', 'SUBSCRIBERKEY']
search_column_name_createddate = ['CreatedDate', 'createddate', 'CREATEDDATE']
search_column_name_lastmodifieddate = ['LastModifiedDate', 'lastmodifieddate', 'LASTMODIFIEDDATE']
search_column_name_campaignname = ['CampaignName', 'campaignname', 'CAMPAIGNNAME']
search_column_name_suppressiontype = ['SuppressionType', 'suppressiontype', 'SUPPRESSIONTYPE']
search_column_name_campaignrunid = ['CampaignRunID', 'campaignrunid', 'CAMPAIGNRUNID']
search_column_name_highlowip = ['HighLowIP', 'highlowip', 'HIGHLOWIP']
search_column_name_sent = ['Sent', 'sent', 'SENT']
search_column_name_sentdate = ['sentdate', 'SENTDATE']
search_column_name_open = ['Open', 'open', 'OPEN']
search_column_name_opendate = ['open date', 'OPEN DATE', 'opendate', 'OPENDATE']
search_column_name_click = ['click', 'CLICK']
search_column_name_clickdate = ['clickdate', 'CLICKDATE']
search_column_name_hardbounce = ['Hardbounce', 'HARDBOUNCE']
search_column_name_softbounce = ['softbounce', 'SOFTBOUNCE']
search_column_name_blockedbounce = ['blockedbounce', 'BLOCKEDBOUNCE']
search_column_name_unsub = ['unsub', 'UNSUB']

# COMMAND ----------

# Step 2: Define function to scrape from the excel the best matching column name with the maximum number of rows
from rapidfuzz import fuzz, utils

'''
regex_pattern = r'[^a-zA-Z]'

def clean_column(column_name_input):
   # Use re.sub to replace the matched keywords with an empty string
  column_name_output = re.sub(regex_pattern, '', column_name_input)
  column_name_output = column_name_output.lower()
  column_name_output = column_name_output.strip()
  return column_name_output
'''

def fuzzy_column_match(df, column_name):
    print(df)  
    #start
    matches = [(col, fuzz.ratio(column_name, col)) for col in df.columns]
    best_match = max(matches, key=lambda x: x[1]) # get the best value from the maximum value of ratio 

    # Count non-null, non-NaN, and non-empty rows in the specified column
    x_count = len(df[df[best_match[0]].notnull() & (df[best_match[0]] != "") & (df[best_match[0]] != " ") ])

    data_to_add = {"Column_Name_Search_Query": column_name, "Dataframe_Column_Name_Found": best_match[0], "Dataframe_Column_Fuzzy_Match_Score": best_match[1], "Dataframe_Column_Row_Count": x_count}
    #print(data_to_add)
    return data_to_add # sending the dictionary back
    #end


# COMMAND ----------

# Step 3: Prepare a new Dictionary to bring the data in Silver Layer

# Create an empty dictionary to store DataFrames in silver layer
silver_dataframe = pd.DataFrame()

for file_name, df in bronze_dataframes.items():  
  print(file_name)

  # Create an empty DataFrame to store matched columns
  silver_df = pd.DataFrame()

  ##############################################################################################################################################################
  # Create dummy dataframe columns

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names
  
  #start
  flag = 0
  for search_col in search_column_name_emailaddress:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe
    
  #print(best_match)
  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1   
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['emailaddress'] = df[max_row_column_name].fillna('')
  else:
    silver_df['emailaddress'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names  
  
  #start
  flag = 0
  for search_col in search_column_name_mobilenumber:
  
    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['mobilenumber'] = df[max_row_column_name].fillna('')
  else:
    silver_df['mobilenumber'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names  
  
  #start
  flag = 0
  for search_col in search_column_name_firstname:
  
    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['firstname'] = df[max_row_column_name].fillna('')
  else:
    silver_df['firstname'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names    
  
  #start
  flag = 0
  for search_col in search_column_name_middlename:
  
    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score] 
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['middlename'] = df[max_row_column_name].fillna('')
  else:
    silver_df['middlename'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names
  
  #start
  flag = 0
  for search_col in search_column_name_lastname:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['lastname'] = df[max_row_column_name].fillna('')
  else:
    silver_df['lastname'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_addr1:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]   
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['addr1'] = df[max_row_column_name].fillna('')
  else:
    silver_df['addr1'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_addr2:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['addr2'] = df[max_row_column_name].fillna('')
  else:
    silver_df['addr2'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_city:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['city'] = df[max_row_column_name].fillna('')
  else:
    silver_df['city'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_state:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['state'] = df[max_row_column_name].fillna('')
  else:
    silver_df['state'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_zip:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['zip'] = df[max_row_column_name].fillna('')
  else:
    silver_df['zip'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_optouturl:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['optouturl'] = df[max_row_column_name].fillna('')
  else:
    silver_df['optouturl'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_var1:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['var1'] = df[max_row_column_name].fillna('')
  else:
    silver_df['var1'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_var2:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['var2'] = df[max_row_column_name].fillna('')
  else:
    silver_df['var2'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_var3:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['var3'] = df[max_row_column_name].fillna('')
  else:
    silver_df['var3'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_var4:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['var4'] = df[max_row_column_name].fillna('')
  else:
    silver_df['var4'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_cid:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['cid'] = df[max_row_column_name].fillna('')
  else:
    silver_df['cid'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_subscriberkey:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['subscriberkey'] = df[max_row_column_name].fillna('')
  else:
    silver_df['subscriberkey'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_createddate:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['createddate'] = df[max_row_column_name].fillna('')
  else:
    silver_df['createddate'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0  
  for search_col in search_column_name_lastmodifieddate:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['lastmodifieddate'] = df[max_row_column_name].fillna('')
  else:
    silver_df['lastmodifieddate'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_campaignname:
  
    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['campaignname'] = df[max_row_column_name].fillna('')
  else:
    silver_df['campaignname'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_suppressiontype:
  
    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['suppressiontype'] = df[max_row_column_name].fillna('')
  else:
    silver_df['suppressiontype'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_campaignrunid:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['campaignrunid'] = df[max_row_column_name].fillna('')
  else:
    silver_df['campaignrunid'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_highlowip:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    silver_df['highlowip'] = df[max_row_column_name].fillna('')    
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    silver_df['highlowip'] = df[max_row_column_name].fillna('')
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end
  
  if flag == 0:
    silver_df['highlowip'] = df[max_row_column_name].fillna('')
  else:
    silver_df['highlowip'] = np.nan
    
  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0    
  for search_col in search_column_name_sent:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    silver_df['sent'] = df[max_row_column_name].fillna('')       
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    silver_df['sent'] = df[max_row_column_name].fillna('')    
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['sent'] = df[max_row_column_name].fillna('')
  else:
    silver_df['sent'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_sentdate:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['sentdate'] = df[max_row_column_name].fillna('')
  else:
    silver_df['sentdate'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_open:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  #print(best_match)
  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['open'] = df[max_row_column_name].fillna('')
  else:
    silver_df['open'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_opendate:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  #print(best_match)
  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['opendate'] = df[max_row_column_name].fillna('')
  else:
    silver_df['opendate'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_click:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  #print(best_match)
  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['click'] = df[max_row_column_name].fillna('')
  else:
    silver_df['click'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_clickdate:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  #print(best_match)
  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['clickdate'] = df[max_row_column_name].fillna('')
  else:
    silver_df['clickdate'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_hardbounce:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  #print(best_match)
  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['hardbounce'] = df[max_row_column_name].fillna('')
  else:
    silver_df['hardbounce'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_softbounce:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  #print(best_match)
  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['softbounce'] = df[max_row_column_name].fillna('')
  else:
    silver_df['softbounce'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_blockedbounce:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  #print(best_match)
  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['blockedbounce'] = df[max_row_column_name].fillna('')
  else:
    silver_df['blockedbounce'] = np.nan

  ##############################################################################################################################################################
  # Create an empty DataFrame to store tuples

  best_match = pd.DataFrame(columns=["Column_Name_Search_Query", "Dataframe_Column_Name_Found", "Dataframe_Column_Fuzzy_Match_Score", "Dataframe_Column_Row_Count"])  # Replace with your column names

  #start
  flag = 0
  for search_col in search_column_name_unsub:

    data_to_add = fuzzy_column_match(df, search_col)

    new_row = pd.DataFrame([data_to_add]) # converting to pandas dataframe

    best_match = pd.concat([best_match, new_row], ignore_index=True) # append to a new dataframe

  #print(best_match)
  # Find the index of the row with the maximum value in Dataframe_Column_Fuzzy_Match_Score
  best_match['Dataframe_Column_Fuzzy_Match_Score'] = best_match['Dataframe_Column_Fuzzy_Match_Score'].astype(float)
  max_index_fuzzy_score = best_match['Dataframe_Column_Fuzzy_Match_Score'].idxmax()
    
  # Find the index of the row with the maximum value in Dataframe_Column_Row_Count
  best_match['Dataframe_Column_Row_Count'] = best_match['Dataframe_Column_Row_Count'].astype(float)
  max_index_row_count = best_match['Dataframe_Column_Row_Count'].idxmax()  

  if max_index_fuzzy_score == max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0 and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    silver_df['unsub'] = df[max_row_column_name].fillna('')    
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Fuzzy_Match_Score'][max_index_fuzzy_score] > 80:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    silver_df['unsub'] = df[max_row_column_name].fillna('')
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_fuzzy_score] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_fuzzy_score]
    flag = 1
  elif max_index_fuzzy_score != max_index_row_count and best_match['Dataframe_Column_Row_Count'][max_index_row_count] > 0:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  else:
    max_row_column_name = best_match['Dataframe_Column_Name_Found'].loc[max_index_row_count]
    flag = 1
  #end

  if flag == 0:
    silver_df['unsub'] = df[max_row_column_name].fillna('')
  else:
    silver_df['unsub'] = np.nan

  ##############################################################################################################################################################

  
  silver_df['file_name'] = file_name
  silver_df['file_name'] = silver_df['file_name'].fillna('')

  ##############################################################################################################################################################

  silver_dataframe = pd.concat([silver_dataframe, silver_df], ignore_index=True)


# COMMAND ----------

#####################################################################################################################################
#Step 0
#####################################################################################################################################
# Create rowid column in dataframe
silver_dataframe['rowid'] = np.arange(silver_dataframe.shape[0])

#####################################################################################################################################
# Step 1
#####################################################################################################################################
# Converting the date column
silver_dataframe['createddate'] = pd.to_datetime(silver_dataframe['createddate'])
# Extract just the date part for non-null values
silver_dataframe['createddate'] = silver_dataframe['createddate'].dt.date
silver_dataframe['createddate'] = silver_dataframe['createddate'].astype(str)

# Converting the date column
silver_dataframe['lastmodifieddate'] = pd.to_datetime(silver_dataframe['lastmodifieddate'])
# Extract just the date part for non-null values
silver_dataframe['lastmodifieddate'] = silver_dataframe['lastmodifieddate'].dt.date
silver_dataframe['lastmodifieddate'] = silver_dataframe['lastmodifieddate'].astype(str)

# Converting the date column
silver_dataframe['sentdate'] = pd.to_datetime(silver_dataframe['sentdate'])
# Extract just the date part for non-null values
silver_dataframe['sentdate'] = silver_dataframe['sentdate'].dt.date
silver_dataframe['sentdate'] = silver_dataframe['sentdate'].astype(str)


# Converting the date column
silver_dataframe['opendate'] = pd.to_datetime(silver_dataframe['opendate'])
# Extract just the date part for non-null values
silver_dataframe['opendate'] = silver_dataframe['opendate'].dt.date
silver_dataframe['opendate'] = silver_dataframe['opendate'].astype(str)

# Converting the date column
silver_dataframe['clickdate'] = pd.to_datetime(silver_dataframe['clickdate'])
# Extract just the date part for non-null values
silver_dataframe['clickdate'] = silver_dataframe['clickdate'].dt.date
silver_dataframe['clickdate'] = silver_dataframe['clickdate'].astype(str)

# Converting the date column
silver_dataframe['hardbounce'] = pd.to_datetime(silver_dataframe['hardbounce'])
# Extract just the date part for non-null values
silver_dataframe['hardbounce'] = silver_dataframe['hardbounce'].dt.date
silver_dataframe['hardbounce'] = silver_dataframe['hardbounce'].astype(str)

# Converting the date column
silver_dataframe['softbounce'] = pd.to_datetime(silver_dataframe['softbounce'])
# Extract just the date part for non-null values
silver_dataframe['softbounce'] = silver_dataframe['softbounce'].dt.date
silver_dataframe['softbounce'] = silver_dataframe['softbounce'].astype(str)

# Converting the date column
silver_dataframe['blockedbounce'] = pd.to_datetime(silver_dataframe['blockedbounce'])
# Extract just the date part for non-null values
silver_dataframe['blockedbounce'] = silver_dataframe['blockedbounce'].dt.date
silver_dataframe['blockedbounce'] = silver_dataframe['blockedbounce'].astype(str)

# Converting the date column
silver_dataframe['unsub'] = pd.to_datetime(silver_dataframe['unsub'])
# Extract just the date part for non-null values
silver_dataframe['unsub'] = silver_dataframe['unsub'].dt.date
silver_dataframe['unsub'] = silver_dataframe['unsub'].astype(str)

#####################################################################################################################################
# Step 2
#####################################################################################################################################
def split_name(row):
    if pd.isnull(row['lastname']) and isinstance(row['firstname'], str) and ' ' in row['firstname'].strip():  # first check for string type before checking for spaces
        parts = row['firstname'].split()
        row['firstname'] = parts[0]  # Take the first part for the first name
        row['lastname'] = ' '.join(parts[1:])  # Combine the remaining parts for the last name
    return row

# Apply the function to each row
silver_dataframe = silver_dataframe.apply(split_name, axis=1)

#####################################################################################################################################
# Step 3
#####################################################################################################################################
# Function to update 'cid' if it's null
def update_cid(row):
    if pd.isnull(row['cid']):
        # Concatenate non-null values from var1 to var4
        row['cid'] = ''.join([str(row[var]) for var in ['var1', 'var2', 'var3', 'var4'] if pd.notnull(row[var])])
    return row

# Apply the function to each row
silver_dataframe = silver_dataframe.apply(update_cid, axis=1)

#####################################################################################################################################
# Step 4
#####################################################################################################################################

#change order of columns
new_order = ['rowid', 'emailaddress', 'mobilenumber', 'firstname', 'middlename', 'lastname', 'addr1', 'addr2', 'city', 'state', 'zip', 'optouturl', 'var1', 'var2', 'var3', 'var4', 'cid', 'subscriberkey', 'createddate', 'lastmodifieddate', 'campaignname', 'suppressiontype', 'campaignrunid', 'highlowip', 'sent', 'sentdate', 'open', 'opendate', 'click', 'clickdate', 'hardbounce', 'softbounce', 'blockedbounce', 'unsub', 'file_name']

# Rearrange the columns
silver_dataframe = silver_dataframe[new_order]

# COMMAND ----------

# MAGIC %md ### Silver Layer - Move to Delta Lake table

# COMMAND ----------

# DBTITLE 1,Create Schema so that errors can be captured
df = spark.createDataFrame(silver_dataframe, schema="rowid BIGINT, emailaddress STRING, mobilenumber STRING, firstname STRING, middlename STRING, lastname STRING, addr1 STRING, addr2 STRING, city STRING, state STRING, zip STRING, optouturl STRING, var1 STRING, var2 STRING, var3 STRING, var4 STRING, cid STRING, subscriberkey STRING, createddate STRING, lastmodifieddate STRING, campaignname STRING, suppressiontype STRING, campaignrunid STRING, highlowip STRING, sent STRING, sentdate STRING, open STRING, opendate STRING, click STRING, clickdate STRING, hardbounce STRING, softbounce STRING, blockedbounce STRING, unsub STRING, file_name STRING")

# COMMAND ----------

df.write.format("delta").mode("append").partitionBy("file_name").saveAsTable("sle_mun_data_silver_layer")

# COMMAND ----------

# MAGIC %md ### Gold Layer - Harmonize the Data

# COMMAND ----------

# MAGIC %sql
# MAGIC
# MAGIC CREATE OR REPLACE table sle_mun_data_gold_layer using delta as
# MAGIC
# MAGIC select 
# MAGIC
# MAGIC a.cid as cid,
# MAGIC upper(a.firstname) as firstname,
# MAGIC upper(a.middlename) as middlename,
# MAGIC upper(a.lastname) as lastname,
# MAGIC a.subscriberkey as subscriberkey,
# MAGIC (case when a.createddate like 'NaT' then null else a.createddate end) as createddate,
# MAGIC (case when a.lastmodifieddate like 'NaT' then null else a.lastmodifieddate end)  as lastmodifieddate,
# MAGIC a.campaignname as campaignname,
# MAGIC a.suppressiontype as suppressiontype,
# MAGIC a.campaignrunid as campaignrunid,
# MAGIC a.highlowip as highlowip,
# MAGIC a.sent as sent,
# MAGIC (case when a.sentdate like 'NaT' then null else a.sentdate end)  as sentdate,
# MAGIC
# MAGIC a.open as open,
# MAGIC (case when a.opendate like 'NaT' then null else a.opendate end)  as opendate,
# MAGIC a.click as click,
# MAGIC (case when a.clickdate like 'NaT' then null else a.clickdate end) as clickdate,
# MAGIC (case when a.hardbounce like 'NaT' then null else a.hardbounce end) as hardbounce,
# MAGIC (case when a.softbounce like 'NaT' then null else a.softbounce end) as softbounce,
# MAGIC (case when a.blockedbounce like 'NaT' then null else a.blockedbounce end)  as blockedbounce,
# MAGIC (case when a.unsub like 'NaT' then null else a.unsub end) as unsub,
# MAGIC a.file_name as sourcefile,
# MAGIC
# MAGIC
# MAGIC from  sle_mun_data_silver_layer a;
# MAGIC optimize sle_mun_data_gold_layer;
